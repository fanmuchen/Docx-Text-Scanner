import os
import re
import json
import docx2txt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins

def standardize_filename(filename):
    match = re.match(r"(\d+)[-_\s]*(.*)", filename)
    if (match):
        number = match.group(1).zfill(3)  # 将数字填充为三位数
        rest = match.group(2).strip()  # 去掉开头的符号和空格
        new_filename = f"{number}-{rest}"
        return new_filename
    return filename

def read_docx(file_path):
    all_text = docx2txt.process(file_path)
    return all_text

def count_keywords(text, keywords):
    counts = {keyword: text.count(keyword) for keyword in keywords}
    return counts

def format_sheet(ws, num_files):
    # 设置第一列宽度为52
    ws.column_dimensions['A'].width = 52
    
    # 设置第二列宽度为12
    ws.column_dimensions['B'].width = 12
    
    # 设置第三列开始的列宽为15
    for col in ws.iter_cols(min_col=3, max_col=ws.max_column):
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 15
    
    # 设置所有单元格自动换行
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    
    # 设置第一列上下居中，靠左对齐
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
            cell.font = Font(name='仿宋_GB2312', size=12, bold=False)
    
    # 设置后面所有列上下居中，左右居中
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            cell.font=Font(name='Times New Roman', size=16)
    
    # 设置列首（第一行）字体为黑体，12号字，不加粗
    for cell in ws[1]:
        cell.font = Font(name='黑体', size=12, bold=False)
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    
    # 添加条件格式，对大于0的单元格标记浅红色背景
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    start_col = get_column_letter(3)  # 第三列
    end_col = get_column_letter(ws.max_column)  # 最右列
    start_row = 3  # 第二行
    end_row = ws.max_row  # 最后一行
    cell_range = f"{start_col}{start_row}:{end_col}{end_row}"
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill))
    
    # 为整张表格添加所有框线
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # 在第一行上面添加一行作为标题
    ws.insert_rows(1)
    
    # 合并标题行单元格
    title_cell = ws.cell(row=1, column=1, value=f'对{num_files}个文件的扫描结果')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    
    # 设置标题行字体为“方正小标宋_GBK”，24号字
    title_cell.font = Font(name='方正小标宋_GBK', size=24)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置所有行高为56
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 56
    
    # 设置打印缩放规则，将所有列打印在一页宽，允许多页高
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # 设置页边距
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)

def main():
    directory = 'files'

    # 从 config.json 文件中读取关键词
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
        keywords = config['keywords']

    data = []

    files = [f for f in os.listdir(directory) if f.endswith('.docx')]
    standardized_files = []

    for filename in files:
        new_filename = standardize_filename(filename)
        old_file_path = os.path.join(directory, filename)
        new_file_path = os.path.join(directory, new_filename)
        
        os.rename(old_file_path, new_file_path)
        standardized_files.append(new_filename)

    standardized_files.sort()

    for filename in standardized_files:
        file_path = os.path.join(directory, filename)
        try:
            all_text = read_docx(file_path)
            word_count = len(all_text)  # 统计全文字数
            keyword_counts = count_keywords(all_text, keywords)
            row_data = {
                '文件名': filename,
                '正文字数': word_count,
                **keyword_counts
            }
            data.append(row_data)
        except Exception as e:
            print(f"Error processing file {filename}: {e}")

    df = pd.DataFrame(data)
    
    # 修改列名
    new_columns = ['文件名', '正文字数'] + [f'“{kw}”\n匹配次数' for kw in keywords]
    df.columns = new_columns
    
    # 保存到 Excel 文件
    df.to_excel('output.xlsx', index=False)
    
    # 调整列宽
    wb = load_workbook('output.xlsx')
    ws = wb.active
    format_sheet(ws, len(standardized_files))
    wb.save('output.xlsx')

    print(f"工作完成，共处理了 {len(standardized_files)} 个文件。")

if __name__ == "__main__":
    main()
